import boto3
import os
import logging
import datetime
import json
import uuid
import mimetypes
import email
import json;
import csv
import openpyxl
import re

from pprint import pprint
from pathlib import Path
import requests
from datetime import timedelta, timezone
import time

LOG_LEVEL = os.environ.get('LOG_LEVEL', 'INFO')
logging.basicConfig(level=LOG_LEVEL,
                    format='{"date": "%(asctime)s", "level": "%(levelname)s", "location": "%(filename)s:%(lineno)d", "message": "%(message)s"}')

def get_secret(key):
    resp =  client.get_secret_value(
        SecretId=key
    )
    return resp['SecretString']

s3_client = boto3.client('s3')
client = boto3.client('secretsmanager', region_name='eu-west-2')
hec_uri = "https://{}/".format(os.environ.get('hec_uri','hec.onboarding.splunk.aws.digital.nhs.uk'))
hec_token = get_secret(os.environ.get('hec_token_key'))
hec_metadata = {"index":"logs_pinnaclevaccstats_prod", "host":"lambda"}

os.environ['TZ'] = 'Europe/London'
time.tzset()


def sendHEC(uri, event, token, metadata=None):
    headers = {
        'Authorization': 'Splunk '+token
        #"X-Splunk-Request-Channel": str(uuid.uuid4())
    }

    r = requests.post("{}/services/collector/raw?host=lambda&source={}&index={}&sourcetype={}".format(uri,hec_metadata['source'],hec_metadata['index'],hec_metadata['sourcetype']), data=json.dumps(event), headers=headers, verify=True if 'https' in uri else False)

    return r.status_code, r.text,

# noinspection PyUnusedLocal
def handler(event, context):

    for record in event['Records']:
        message = json.loads(record['Sns']['Message'])

        logging.warning("Source: {}".format(message['mail']['source']))
        logging.warning("VirusVerdict: {}".format(message['receipt']['virusVerdict']['status']))

        mail_bucket = message['receipt']['action']['bucketName']
        mail_objectKeyPrefix = "{}/".format(message['receipt']['action']['objectKeyPrefix']) if 'objectKeyPrefix' in message['receipt']['action'] and  message['receipt']['action']['objectKeyPrefix'] != '' else ""
        mail_objectKey = message['receipt']['action']['objectKey']
        logging.warning("S3Path: s3://{}{}/{}".format(mail_bucket,mail_objectKeyPrefix,mail_objectKey))

        tmpkey = mail_objectKey.replace('/', '')
        download_path = '/tmp/{}{}'.format(uuid.uuid4(), tmpkey)
        s3_client.download_file(mail_bucket, "{}{}".format(mail_objectKeyPrefix,mail_objectKey), download_path)
        logging.warning(download_path)
        with open(download_path) as fp:
            filename = ""
            msg = email.message_from_file(fp)
            counter = 1
            for part in msg.walk():

                # multipart/* are just containers
                if part.get_content_maintype() == 'multipart':
                    filename = ""
                    print("This part is multipart..")

                    for subpart in part.walk():
                        try:
                            #filename = subpart.get_filename()
                            filename = subpart.get_param('name') or subpart.get_filename()
                            print("Filename={}".format(filename))
                        except:
                            filename = ""
                        if filename!=None:
                            print("BREAK")
                            part = subpart
                            break
                    if filename != "":
                        print("BREAK2")

                if filename=="":
                    filename = "{}".format(part.get_filename())
                if not filename:
                    ext = mimetypes.guess_extension(part.get_content_type())
                    if not ext:
                        # Use a generic bag-of-bits extension
                        ext = '.bin'
                    filename = 'part-%03d%s' % (counter, ext)
                counter += 1
                splitfilename = os.path.splitext(filename)
                pprint(splitfilename)
                if splitfilename[1] == ".csv":
                    hec_metadata['source'] = filename
                    hec_metadata['sourcetype'] = 'itoc:vacc:pinnacle:stats'

                    with open(os.path.join("/tmp", filename), 'wb') as fp:
                        file_contents = part.get_payload(decode=True)
                        fp.write(file_contents)

                    rcv_file = open('/tmp/{}'.format(filename))

                    csv.register_dialect('comma', delimiter=',')
                    reader = csv.reader(rcv_file, dialect='comma')

                    wb = openpyxl.Workbook()
                    sheet = wb.active

                    for row in reader:
                        sheet.append(row)

                    events = {}
                    i = 0

                    for rowNum in range(2, sheet.max_row + 1):
                        # Region
                        regionText = sheet.cell(row=rowNum, column=1).value

                        region = regionText if regionText!="" else "None"

                        events[i] = {
                            'region': region,
                            'timestamp': sheet.cell(row=rowNum, column=2).value + "+01:00"
                        }

                        for colNum in range(3, sheet.max_column + 1):
                            section = sheet.cell(row=1, column=colNum).value
                            sectionSafe = section.replace("%","perc").replace(" ","_")

                            cellVal = sheet.cell(row=rowNum,column=colNum).value
                            if cellVal!="":
                                events[i][sectionSafe] = cellVal

                        i += 1
                    if events!={}:
                        for event in events.values():
                            print(event)
                            r_status, r_text = sendHEC(uri=hec_uri, event=event, token=hec_token, metadata=hec_metadata)
                            print("HEC Status={} text=\"{}\"".format(r_status, r_text))
                    return True
    return {
        'statusCode': 200,
        'body': 'Completed'
    }




